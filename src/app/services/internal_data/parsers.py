"""Parsers for Internal Data sources (CSV, XLSX, XML)."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Tuple


EXPECTED_COLUMNS = {
    "internal_sku",
    "marketplace_code",
    "marketplace_sku",
    "marketplace_item_id",
    "rrp",
    "cost",
    "lifecycle_status",
    "attributes_json",
}


@dataclass
class ParsedRow:
    internal_sku: str
    name: str | None
    lifecycle_status: str | None
    attributes: Dict[str, Any] | None
    identifiers: List[Dict[str, Any]]
    price: Dict[str, Any] | None
    cost: Dict[str, Any] | None


def _normalize_row(raw: Dict[str, str]) -> ParsedRow:
    internal_sku = (raw.get("internal_sku") or "").strip()
    if not internal_sku:
        raise ValueError("internal_sku is required")

    marketplace_code = (raw.get("marketplace_code") or "").strip()
    marketplace_sku = (raw.get("marketplace_sku") or "").strip() or None
    marketplace_item_id = (raw.get("marketplace_item_id") or "").strip() or None

    identifiers: List[Dict[str, Any]] = []
    if marketplace_code:
        identifiers.append(
            {
                "marketplace_code": marketplace_code,
                "marketplace_sku": marketplace_sku,
                "marketplace_item_id": marketplace_item_id,
                "extra_identifiers": None,
            }
        )

    def _parse_decimal(value: str | None) -> Any:
        if value is None:
            return None
        v = value.replace(" ", "").replace(",", ".")
        if not v:
            return None
        try:
            return float(v)
        except ValueError:
            return None

    rrp_val = _parse_decimal(raw.get("rrp"))
    cost_val = _parse_decimal(raw.get("cost"))

    price = {"currency": "RUB", "rrp": rrp_val, "rrp_promo": None, "extra": None} if rrp_val is not None else None
    cost = {"currency": "RUB", "cost": cost_val, "extra": None} if cost_val is not None else None

    lifecycle_status = (raw.get("lifecycle_status") or "").strip() or None

    attributes = None
    attrs_raw = (raw.get("attributes_json") or "").strip()
    if attrs_raw:
        import json

        try:
            parsed = json.loads(attrs_raw)
            if isinstance(parsed, dict):
                attributes = parsed
        except json.JSONDecodeError:
            attributes = None

    # Any extra columns (not in EXPECTED_COLUMNS) are merged into attributes
    extra_attrs: Dict[str, Any] = {}
    for key, value in raw.items():
        if key not in EXPECTED_COLUMNS and value not in (None, ""):
            extra_attrs[key] = value
    if extra_attrs:
        attributes = {**(attributes or {}), **extra_attrs}

    return ParsedRow(
        internal_sku=internal_sku,
        name=None,
        lifecycle_status=lifecycle_status,
        attributes=attributes,
        identifiers=identifiers,
        price=price,
        cost=cost,
    )


def parse_csv(path: str) -> Iterable[Dict[str, Any]]:
    """Yield normalized rows from CSV file."""
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for raw in reader:
            if not raw:
                continue
            try:
                parsed = _normalize_row({k.strip(): (v or "").strip() for k, v in raw.items()})
            except ValueError:
                continue
            yield {
                "internal_sku": parsed.internal_sku,
                "name": parsed.name,
                "lifecycle_status": parsed.lifecycle_status,
                "attributes": parsed.attributes,
                "identifiers": parsed.identifiers,
                "price": parsed.price,
                "cost": parsed.cost,
            }


def parse_xlsx(path: str) -> Iterable[Dict[str, Any]]:
    """Yield normalized rows from XLSX file."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []

    header = [str(h).strip() if h is not None else "" for h in header]
    indices = {name: idx for idx, name in enumerate(header) if name}

    def get_val(row, col_name: str) -> str | None:
        idx = indices.get(col_name)
        if idx is None or idx >= len(row):
            return None
        value = row[idx]
        return "" if value is None else str(value)

    out: List[Dict[str, Any]] = []
    for row in rows_iter:
        raw: Dict[str, str] = {}
        for col in EXPECTED_COLUMNS:
            raw[col] = get_val(row, col) or ""
        # include any other columns for attributes
        for name, idx in indices.items():
            if name in EXPECTED_COLUMNS:
                continue
            raw[name] = "" if idx >= len(row) or row[idx] is None else str(row[idx])
        try:
            parsed = _normalize_row(raw)
        except ValueError:
            continue
        out.append(
            {
                "internal_sku": parsed.internal_sku,
                "name": parsed.name,
                "lifecycle_status": parsed.lifecycle_status,
                "attributes": parsed.attributes,
                "identifiers": parsed.identifiers,
                "price": parsed.price,
                "cost": parsed.cost,
            }
        )
    return out


def parse_xml(path: str) -> List[Dict[str, Any]]:
    """Parse XML file in legacy RRP format into normalized Internal Data rows.

    Expected structure:

      <items>
        <item article="SKU" stock="123" price="84" [barcode="..."] />
        ...
      </items>

    The `article` attribute is treated as the internal SKU (preserved as-is),
    and `price` is mapped to RRP in RUB. Stock and barcode are placed into
    `attributes` for potential downstream use.
    """
    import xml.etree.ElementTree as ET

    rows: List[Dict[str, Any]] = []
    seen_item = False

    try:
        for _, elem in ET.iterparse(path, events=("end",)):
            if elem.tag != "item":
                continue

            seen_item = True

            raw_sku = (elem.attrib.get("article") or "").strip()
            raw_price = (elem.attrib.get("price") or "").strip()
            raw_qty = (elem.attrib.get("stock") or "").strip()
            raw_barcode = (elem.attrib.get("barcode") or "").strip()

            if not raw_sku or not raw_price:
                # Missing required fields – skip this item
                elem.clear()
                continue

            # Use raw_sku as-is, no automatic normalization
            # Transforms from mapping_json will be applied if mapping is used
            sku = raw_sku

            # Parse price as float (RUB)
            price_str = raw_price.replace(" ", "").replace(",", ".")
            try:
                price_val = float(price_str)
            except ValueError:
                elem.clear()
                continue

            qty_val = None
            if raw_qty:
                try:
                    qty_val = int(raw_qty)
                except Exception:
                    qty_val = None

            attrs: Dict[str, Any] = {}
            if qty_val is not None:
                attrs["stock"] = qty_val
            elif raw_qty:
                # preserve raw value if it was non-empty but not an int
                attrs["stock_raw"] = raw_qty
            if raw_barcode:
                attrs["barcode"] = raw_barcode

            rows.append(
                {
                    "internal_sku": sku,
                    "name": None,
                    "lifecycle_status": None,
                    "attributes": attrs or None,
                    "identifiers": [],
                    "price": {"currency": "RUB", "rrp": price_val, "rrp_promo": None, "extra": None},
                    "cost": None,
                }
            )

            # Help the GC when parsing large XML files
            elem.clear()
    except ET.ParseError as exc:  # malformed XML
        raise ValueError(f"Invalid XML: {exc}") from exc

    if not seen_item:
        # No <item> elements at all – most likely wrong XML structure
        raise ValueError('Invalid Internal Data XML: expected <items><item article="..." price="..."/></items>')

    if not rows:
        # We saw <item> but none had valid article/price
        raise ValueError("No valid <item> elements with 'article' and numeric 'price' were found in XML")

    return rows


def introspect_csv(path: str, max_rows: int = 3) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Return headers and first few raw rows from a CSV file."""
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        headers = [h.strip() for h in (reader.fieldnames or []) if h and h.strip()]
        rows: List[Dict[str, Any]] = []
        for idx, raw in enumerate(reader):
            if idx >= max_rows:
                break
            if not raw:
                continue
            # Keep raw values as-is but strip header names
            normalized: Dict[str, Any] = {}
            for k, v in raw.items():
                if k is None:
                    continue
                key = k.strip()
                if not key:
                    continue
                normalized[key] = v
            if normalized:
                rows.append(normalized)
        return headers, rows


def introspect_xlsx(path: str, max_rows: int = 3) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Return headers and first few raw rows from an XLSX file."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], []

    headers = [str(h).strip() if h is not None else "" for h in header]
    headers = [h for h in headers if h]
    indices = {name: idx for idx, name in enumerate(header) if name}

    preview_rows: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows_iter):
        if idx >= max_rows:
            break
        raw: Dict[str, Any] = {}
        for name, col_idx in indices.items():
            if col_idx >= len(row):
                continue
            value = row[col_idx]
            raw[str(name).strip()] = value
        if raw:
            preview_rows.append(raw)

    return headers, preview_rows


def introspect_xml(
    path: str,
    *,
    max_items: int = 20,
    preview_items: int = 3,
) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Return attribute names and first few <item> attribute maps from an XML file."""
    import xml.etree.ElementTree as ET

    attribute_names_set = set()
    preview: List[Dict[str, Any]] = []
    seen = 0

    try:
        for _, elem in ET.iterparse(path, events=("end",)):
            if elem.tag != "item":
                continue
            seen += 1
            attrs = dict(elem.attrib)
            attribute_names_set.update(attrs.keys())
            if len(preview) < preview_items:
                preview.append(attrs)
            elem.clear()
            if seen >= max_items:
                break
    except ET.ParseError as exc:
        raise ValueError(f"Invalid XML: {exc}") from exc

    return sorted(attribute_names_set), preview


def iter_rows_csv(path: str) -> Iterable[Dict[str, Any]]:
    """Iterate over raw CSV rows as dicts with header keys.

    Keys match header strings exactly, as in introspect_csv/sample_rows.
    """
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for raw in reader:
            if not raw:
                continue
            row: Dict[str, Any] = {}
            for k, v in raw.items():
                if k is None:
                    continue
                key = k.strip()
                if not key:
                    continue
                row[key] = v
            if row:
                yield row


def iter_rows_xlsx(path: str) -> Iterable[Dict[str, Any]]:
    """Iterate over raw XLSX rows as dicts with header keys."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return

    headers = [str(h).strip() if h is not None else "" for h in header]
    headers = [h for h in headers if h]
    indices = {name: idx for idx, name in enumerate(header) if name}

    for row in rows_iter:
        raw: Dict[str, Any] = {}
        for name, col_idx in indices.items():
            if col_idx >= len(row):
                continue
            value = row[col_idx]
            raw[str(name).strip()] = value
        if raw:
            yield raw


def iter_items_xml(path: str) -> Iterable[Dict[str, Any]]:
    """Iterate over XML <item> attributes as dicts with '@attr' keys."""
    import xml.etree.ElementTree as ET

    for _, elem in ET.iterparse(path, events=("end",)):
        if elem.tag != "item":
            continue
        attrs = {f"@{k}": v for k, v in elem.attrib.items()}
        if attrs:
            yield attrs
        elem.clear()

