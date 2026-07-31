"""Parse WB daily sales-funnel reports from XLSX, CSV, or a ZIP containing one report."""

from __future__ import annotations

import csv
import hashlib
import io
import os
import re
import tempfile
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from app.services.wb_funnel_quality import evaluate_daily_quality


class WBFunnelReportError(ValueError):
    pass


@dataclass(frozen=True)
class WBFunnelReportRow:
    row_number: int
    nm_id: int
    stat_date: date
    vendor_code: str | None
    product_name: str | None
    is_deleted: bool
    impressions: int
    card_clicks: int
    reported_ctr: Decimal | None
    quality_status: str
    quality_flags: tuple[str, ...]
    source_payload: dict[str, Any]


@dataclass(frozen=True)
class ParsedWBFunnelReport:
    source_type: str
    content_sha256: str
    rows: tuple[WBFunnelReportRow, ...]


def _normalize_header(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).replace("ё", "е").lower()
    return re.sub(r"\s+", " ", text).strip(" \t\r\n.,:;_")


ALIASES = {
    "vendor_code": {"артикул продавца", "артикул поставщика"},
    "nm_id": {"артикул wb", "артикул вб", "nm id", "nmid"},
    "product_name": {"название", "наименование"},
    "is_deleted": {"удаленный товар", "удалённый товар"},
    "stat_date": {"дата", "день"},
    "impressions": {"показы", "количество показов"},
    "reported_ctr": {"ctr", "ctr, %", "ctr %"},
    "card_clicks": {"переходы в карточку", "перешли в карточку товара", "открытия карточки"},
}
NORMALIZED_ALIASES = {
    field: {_normalize_header(alias) for alias in aliases} for field, aliases in ALIASES.items()
}
REQUIRED_FIELDS = {"nm_id", "stat_date", "impressions", "card_clicks"}


def _resolve_columns(headers: Iterable[Any]) -> dict[str, int]:
    normalized = [_normalize_header(header) for header in headers]
    resolved: dict[str, int] = {}
    for field, aliases in NORMALIZED_ALIASES.items():
        for index, header in enumerate(normalized):
            if header in aliases:
                resolved[field] = index
                break
    return resolved


def _decimal(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).replace("\u00a0", "").replace(" ", "").replace(",", ".").replace("%", "")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise WBFunnelReportError(f"Invalid numeric value: {value!r}") from exc


def _integer(value: Any, field: str) -> int:
    parsed = _decimal(value)
    if parsed is None or parsed != parsed.to_integral_value():
        raise WBFunnelReportError(f"{field} must be an integer, got {value!r}")
    result = int(parsed)
    if result < 0:
        raise WBFunnelReportError(f"{field} must be non-negative, got {value!r}")
    return result


def _date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    raise WBFunnelReportError(f"Invalid date: {value!r}")


def _is_deleted(value: Any) -> bool:
    return _normalize_header(value) in {"да", "yes", "true", "1", "удален", "удалён"}


def _json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _build_row(row_number: int, headers: list[Any], values: list[Any], columns: dict[str, int]) -> WBFunnelReportRow:
    def get(field: str) -> Any:
        index = columns.get(field)
        return values[index] if index is not None and index < len(values) else None

    nm_id = _integer(get("nm_id"), "nm_id")
    stat_date = _date(get("stat_date"))
    impressions = _integer(get("impressions"), "impressions")
    card_clicks = _integer(get("card_clicks"), "card_clicks")
    reported_ctr = _decimal(get("reported_ctr"))
    deleted = _is_deleted(get("is_deleted"))
    quality_status, quality_flags = evaluate_daily_quality(
        impressions=impressions,
        card_clicks=card_clicks,
        reported_ctr=reported_ctr,
        is_deleted=deleted,
    )
    payload = {
        str(header).strip(): _json_value(values[index] if index < len(values) else None)
        for index, header in enumerate(headers)
        if header is not None and str(header).strip()
    }
    return WBFunnelReportRow(
        row_number=row_number,
        nm_id=nm_id,
        stat_date=stat_date,
        vendor_code=str(get("vendor_code")).strip() if get("vendor_code") not in (None, "") else None,
        product_name=str(get("product_name")).strip() if get("product_name") not in (None, "") else None,
        is_deleted=deleted,
        impressions=impressions,
        card_clicks=card_clicks,
        reported_ctr=reported_ctr,
        quality_status=quality_status,
        quality_flags=tuple(quality_flags),
        source_payload=payload,
    )


def _rows_from_matrix(matrix: list[list[Any]]) -> tuple[WBFunnelReportRow, ...]:
    header_index = None
    columns: dict[str, int] = {}
    for index, candidate in enumerate(matrix[:10]):
        candidate_columns = _resolve_columns(candidate)
        if REQUIRED_FIELDS.issubset(candidate_columns):
            header_index = index
            columns = candidate_columns
            break
    if header_index is None:
        raise WBFunnelReportError("Required columns not found: Артикул WB, Дата, Показы, Переходы в карточку")
    headers = matrix[header_index]
    rows: list[WBFunnelReportRow] = []
    for index, values in enumerate(matrix[header_index + 1 :], start=header_index + 2):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append(_build_row(index, headers, values, columns))
    if not rows:
        raise WBFunnelReportError("Report contains no data rows")
    return tuple(rows)


def _parse_xlsx(path: str) -> tuple[WBFunnelReportRow, ...]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = next((name for name in workbook.sheetnames if _normalize_header(name) == "товары"), None)
        if sheet_name is None:
            raise WBFunnelReportError("XLSX sheet 'Товары' not found")
        worksheet = workbook[sheet_name]
        # Some WB exports declare an incorrect A:A worksheet dimension even
        # though data spans many columns. Recalculate it in read-only mode.
        worksheet.reset_dimensions()
        matrix = [list(row) for row in worksheet.iter_rows(values_only=True)]
        return _rows_from_matrix(matrix)
    finally:
        workbook.close()


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise WBFunnelReportError("CSV encoding must be UTF-8 or Windows-1251")


def _parse_csv(content: bytes) -> tuple[WBFunnelReportRow, ...]:
    text = _decode_csv(content)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";"
    matrix = [list(row) for row in csv.reader(io.StringIO(text), delimiter=delimiter)]
    return _rows_from_matrix(matrix)


def parse_wb_funnel_report(path: str, *, max_uncompressed_bytes: int = 50 * 1024 * 1024) -> ParsedWBFunnelReport:
    suffix = Path(path).suffix.lower()
    if suffix == ".xlsx":
        content = Path(path).read_bytes()
        return ParsedWBFunnelReport("xlsx", hashlib.sha256(content).hexdigest(), _parse_xlsx(path))
    if suffix == ".csv":
        content = Path(path).read_bytes()
        return ParsedWBFunnelReport("csv", hashlib.sha256(content).hexdigest(), _parse_csv(content))
    if suffix != ".zip":
        raise WBFunnelReportError("Supported formats: XLSX, CSV, ZIP")

    try:
        with zipfile.ZipFile(path) as archive:
            candidates = [
                entry for entry in archive.infolist()
                if not entry.is_dir() and Path(entry.filename).suffix.lower() in {".xlsx", ".csv"}
            ]
            if len(candidates) != 1:
                raise WBFunnelReportError("ZIP must contain exactly one XLSX or CSV report")
            entry = candidates[0]
            if Path(entry.filename).is_absolute() or ".." in Path(entry.filename).parts:
                raise WBFunnelReportError("Unsafe ZIP entry path")
            if entry.file_size > max_uncompressed_bytes:
                raise WBFunnelReportError("Uncompressed report is too large")
            if entry.compress_size > 0 and entry.file_size / entry.compress_size > 200:
                raise WBFunnelReportError("Suspicious ZIP compression ratio")
            content = archive.read(entry)
            digest = hashlib.sha256(content).hexdigest()
            if Path(entry.filename).suffix.lower() == ".csv":
                rows = _parse_csv(content)
            else:
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
                    temp_file.write(content)
                    temp_path = temp_file.name
                try:
                    rows = _parse_xlsx(temp_path)
                finally:
                    os.unlink(temp_path)
            return ParsedWBFunnelReport("zip_xlsx" if entry.filename.lower().endswith(".xlsx") else "zip_csv", digest, rows)
    except zipfile.BadZipFile as exc:
        raise WBFunnelReportError("Invalid ZIP archive") from exc
